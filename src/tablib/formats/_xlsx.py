""" Tablib - XLSX Support.
"""

import re
from io import BytesIO

from openpyxl.reader.excel import ExcelReader, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

import tablib

INVALID_TITLE_REGEX = re.compile(r'[\\*?:/\[\]]')


def safe_xlsx_sheet_title(s, replace="-"):
    return re.sub(INVALID_TITLE_REGEX, replace, s)[:31]


class XLSXFormat:
    title = 'xlsx'
    extensions = ('xlsx',)

    @classmethod
    def detect(cls, stream):
        """Returns True if given stream is a readable excel file."""
        try:
            # No need to fully load the file, it should be enough to be able to
            # read the manifest.
            reader = ExcelReader(stream, read_only=False)
            reader.read_manifest()
            return True
        except Exception:
            return False

    @classmethod
    def export_set(cls, dataset, freeze_panes=True, invalid_char_subst="-", escape=False):
        """Returns XLSX representation of Dataset.

        If ``freeze_panes`` is True, Export will freeze panes only after first line.

        If ``dataset.title`` contains characters which are
        considered invalid for an XLSX file sheet name
        (https://web.archive.org/web/20230323081941/https://www.excelcodex.com/2012/06/worksheets-naming-conventions/),
        they will be replaced with ``invalid_char_subst``.

        If ``escape`` is True, formulae will have the leading '=' character removed.
        This is a security measure to prevent formulae from executing by default
        in exported XLSX files.
        """
        wb = Workbook()
        ws = wb.worksheets[0]

        ws.title = (
            safe_xlsx_sheet_title(dataset.title, invalid_char_subst)
            if dataset.title else 'Tablib Dataset'
        )

        cls.dset_sheet(dataset, ws, freeze_panes=freeze_panes, escape=escape)

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    @classmethod
    def export_book(cls, databook, freeze_panes=True, invalid_char_subst="-", escape=False):
        """Returns XLSX representation of DataBook.
        See export_set().
        """

        wb = Workbook()
        for sheet in wb.worksheets:
            wb.remove(sheet)
        for i, dset in enumerate(databook._datasets):
            ws = wb.create_sheet()
            ws.title = (
                safe_xlsx_sheet_title(dset.title, invalid_char_subst)
                if dset.title else f"Sheet{i}"
            )

            cls.dset_sheet(dset, ws, freeze_panes=freeze_panes, escape=escape)

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    @classmethod
    def import_sheet(cls, dset, sheet, headers=True, skip_lines=0):
        """Populates dataset with sheet."""

        dset.title = sheet.title

        for i, row in enumerate(sheet.rows):
            if i < skip_lines:
                continue
            row_vals = [c.value for c in row]
            if i == skip_lines and headers:
                dset.headers = row_vals
            else:
                if i > skip_lines and len(row_vals) < dset.width:
                    row_vals += [''] * (dset.width - len(row_vals))
                dset.append(row_vals)

    @classmethod
    def import_set(cls, dset, in_stream, headers=True, read_only=True, skip_lines=0):
        """Returns databook from XLS stream."""

        dset.wipe()

        xls_book = load_workbook(in_stream, read_only=read_only, data_only=True)
        sheet = xls_book.active
        cls.import_sheet(dset, sheet, headers, skip_lines)

    @classmethod
    def import_book(cls, dbook, in_stream, headers=True, read_only=True):
        """Returns databook from XLS stream."""

        dbook.wipe()

        xls_book = load_workbook(in_stream, read_only=read_only, data_only=True)

        for sheet in xls_book.worksheets:
            dset = tablib.Dataset()
            cls.import_sheet(dset, sheet, headers)
            dbook.add_sheet(dset)

    @classmethod
    def dset_sheet(cls, dataset, ws, freeze_panes=True, escape=False):
        """Completes given worksheet from given Dataset."""
        _package = dataset._package(dicts=False)

        for i, sep in enumerate(dataset._separators):
            _offset = i
            _package.insert((sep[0] + _offset), (sep[1],))

        bold = Font(bold=True)
        wrap_text = Alignment(wrap_text=True)

        for i, row in enumerate(_package):
            row_number = i + 1
            for j, col in enumerate(row):
                col_idx = get_column_letter(j + 1)
                cell = ws[f'{col_idx}{row_number}']

                # bold headers
                if (row_number == 1) and dataset.headers:
                    cell.font = bold
                    if freeze_panes:
                        #  Export Freeze only after first Line
                        ws.freeze_panes = 'A2'

                # bold separators
                elif len(row) < dataset.width:
                    cell.font = bold

                # wrap the rest
                else:
                    if '\n' in str(col):
                        cell.alignment = wrap_text

                try:
                    cell.value = col
                except ValueError:
                    cell.value = str(col)

                if escape and cell.data_type == 'f' and cell.value.startswith('='):
                    cell.value = cell.value.replace("=", "")
