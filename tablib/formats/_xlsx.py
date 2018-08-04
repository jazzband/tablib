# -*- coding: utf-8 -*-

""" Tablib - XLSX Support.
"""

import sys


if sys.version_info[0] > 2:
    from io import BytesIO
else:
    from cStringIO import StringIO as BytesIO

import openpyxl
import tablib

Workbook = openpyxl.workbook.Workbook
ExcelWriter = openpyxl.writer.excel.ExcelWriter
get_column_letter = openpyxl.utils.get_column_letter

from tablib.compat import unicode


title = 'xlsx'
extensions = ('xlsx',)


def detect(stream):
    """Returns True if given stream is a readable excel file."""
    try:
        openpyxl.reader.excel.load_workbook(stream)
        return True
    except openpyxl.shared.exc.InvalidFileException:
        pass

def export_set(dataset, freeze_panes=True):
    """Returns XLSX representation of Dataset."""

    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = dataset.title if dataset.title else 'Tablib Dataset'

    dset_sheet(dataset, ws, freeze_panes=freeze_panes)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_book(databook, freeze_panes=True):
    """Returns XLSX representation of DataBook."""

    wb = Workbook()
    for sheet in wb.worksheets:
        wb.remove_sheet(sheet)
    for i, dset in enumerate(databook._datasets):
        ws = wb.create_sheet()
        ws.title = dset.title if dset.title else 'Sheet%s' % (i)

        dset_sheet(dset, ws, freeze_panes=freeze_panes)


    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def import_set(dset, in_stream, headers=True):
    """Returns databook from XLS stream."""

    dset.wipe()

    xls_book = openpyxl.reader.excel.load_workbook(BytesIO(in_stream))
    sheet = xls_book.get_active_sheet()

    dset.title = sheet.title

    for i, row in enumerate(sheet.rows):
        row_vals = [c.value for c in row]
        if (i == 0) and (headers):
            dset.headers = row_vals
        else:
            dset.append(row_vals)


def import_book(dbook, in_stream, headers=True):
    """Returns databook from XLS stream."""

    dbook.wipe()

    xls_book = openpyxl.reader.excel.load_workbook(BytesIO(in_stream))

    for sheet in xls_book.worksheets:
        data = tablib.Dataset()
        data.title = sheet.title

        for i, row in enumerate(sheet.rows):
            row_vals = [c.value for c in row]
            if (i == 0) and (headers):
                data.headers = row_vals
            else:
                data.append(row_vals)

        dbook.add_sheet(data)


def dset_sheet(dataset, ws, freeze_panes=True):
    """Completes given worksheet from given Dataset."""
    _package = dataset._package(dicts=False)

    for i, sep in enumerate(dataset._separators):
        _offset = i
        _package.insert((sep[0] + _offset), (sep[1],))

    bold = openpyxl.styles.Font(bold=True)
    wrap_text = openpyxl.styles.Alignment(wrap_text=True)

    for i, row in enumerate(_package):
        row_number = i + 1
        for j, col in enumerate(row):
            col_idx = get_column_letter(j + 1)
            cell = ws.cell('%s%s' % (col_idx, row_number))

            # bold headers
            if (row_number == 1) and dataset.headers:
                # cell.value = unicode('%s' % col, errors='ignore')
                cell.value = unicode(col)
                cell.font = bold
                if freeze_panes:
                    #  Export Freeze only after first Line
                    ws.freeze_panes = 'A2'
                    
            # bold separators
            elif len(row) < dataset.width:
                cell.value = unicode('%s' % col, errors='ignore')
                cell.font = bold

            # wrap the rest
            else:
                try:
                    if '\n' in col:
                        cell.value = unicode('%s' % col, errors='ignore')
                        cell.alignment = wrap_text
                    else:
                        cell.value = unicode('%s' % col, errors='ignore')
                except TypeError:
                    cell.value = unicode(col)


