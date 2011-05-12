# file openpyxl/tests/test_worksheet.py

# Copyright (c) 2010 openpyxl
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# @license: http://www.opensource.org/licenses/mit-license.php
# @author: Eric Gazoni

# 3rd party imports
from nose.tools import eq_, raises, assert_raises

# package imports
from openpyxl.workbook import Workbook
from openpyxl.worksheet import Worksheet, Relationship, flatten
from openpyxl.cell import Cell
from openpyxl.shared.exc import CellCoordinatesException, \
        SheetTitleException, InsufficientCoordinatesException, \
        NamedRangeException


class TestWorksheet():

    @classmethod
    def setup_class(cls):
        cls.wb = Workbook()

    def test_new_worksheet(self):
        ws = Worksheet(self.wb)
        eq_(self.wb, ws._parent)

    def test_new_sheet_name(self):
        self.wb.worksheets = []
        ws = Worksheet(self.wb, title = '')
        eq_(repr(ws), '<Worksheet "Sheet1">')

    def test_get_cell(self):
        ws = Worksheet(self.wb)
        cell = ws.cell('A1')
        eq_(cell.get_coordinate(), 'A1')

    @raises(SheetTitleException)
    def test_set_bad_title(self):
        Worksheet(self.wb, 'X' * 50)

    def test_set_bad_title_character(self):
        assert_raises(SheetTitleException, Worksheet, self.wb, '[')
        assert_raises(SheetTitleException, Worksheet, self.wb, ']')
        assert_raises(SheetTitleException, Worksheet, self.wb, '*')
        assert_raises(SheetTitleException, Worksheet, self.wb, ':')
        assert_raises(SheetTitleException, Worksheet, self.wb, '?')
        assert_raises(SheetTitleException, Worksheet, self.wb, '/')
        assert_raises(SheetTitleException, Worksheet, self.wb, '\\')

    def test_worksheet_dimension(self):
        ws = Worksheet(self.wb)
        eq_('A1:A1', ws.calculate_dimension())
        ws.cell('B12').value = 'AAA'
        eq_('A1:B12', ws.calculate_dimension())

    def test_worksheet_range(self):
        ws = Worksheet(self.wb)
        xlrange = ws.range('A1:C4')
        assert isinstance(xlrange, tuple)
        eq_(4, len(xlrange))
        eq_(3, len(xlrange[0]))

    def test_worksheet_named_range(self):
        ws = Worksheet(self.wb)
        self.wb.create_named_range('test_range', ws, 'C5')
        xlrange = ws.range('test_range')
        assert isinstance(xlrange, Cell)
        eq_(5, xlrange.row)

    @raises(NamedRangeException)
    def test_bad_named_range(self):
        ws = Worksheet(self.wb)
        ws.range('bad_range')

    @raises(NamedRangeException)
    def test_named_range_wrong_sheet(self):
        ws1 = Worksheet(self.wb)
        ws2 = Worksheet(self.wb)
        self.wb.create_named_range('wrong_sheet_range', ws1, 'C5')
        ws2.range('wrong_sheet_range')

    def test_cell_offset(self):
        ws = Worksheet(self.wb)
        eq_('C17', ws.cell('B15').offset(2, 1).get_coordinate())

    def test_range_offset(self):
        ws = Worksheet(self.wb)
        xlrange = ws.range('A1:C4', 1, 3)
        assert isinstance(xlrange, tuple)
        eq_(4, len(xlrange))
        eq_(3, len(xlrange[0]))
        eq_('D2', xlrange[0][0].get_coordinate())

    def test_cell_alternate_coordinates(self):
        ws = Worksheet(self.wb)
        cell = ws.cell(row = 8, column = 4)
        eq_('E9', cell.get_coordinate())

    @raises(InsufficientCoordinatesException)
    def test_cell_insufficient_coordinates(self):
        ws = Worksheet(self.wb)
        cell = ws.cell(row = 8)

    def test_cell_range_name(self):
        ws = Worksheet(self.wb)
        self.wb.create_named_range('test_range_single', ws, 'B12')
        assert_raises(CellCoordinatesException, ws.cell, 'test_range_single')
        c_range_name = ws.range('test_range_single')
        c_range_coord = ws.range('B12')
        c_cell = ws.cell('B12')
        eq_(c_range_coord, c_range_name)
        eq_(c_range_coord, c_cell)

    def test_garbage_collect(self):
        ws = Worksheet(self.wb)
        ws.cell('A1').value = ''
        ws.cell('B2').value = '0'
        ws.cell('C4').value = 0
        ws.garbage_collect()
        eq_(ws.get_cell_collection(), [ws.cell('B2'), ws.cell('C4')])

    def test_hyperlink_relationships(self):
        ws = Worksheet(self.wb)
        eq_(len(ws.relationships), 0)

        ws.cell('A1').hyperlink = "http://test.com"
        eq_(len(ws.relationships), 1)
        eq_("rId1", ws.cell('A1').hyperlink_rel_id)
        eq_("rId1", ws.relationships[0].id)
        eq_("http://test.com", ws.relationships[0].target)
        eq_("External", ws.relationships[0].target_mode)

        ws.cell('A2').hyperlink = "http://test2.com"
        eq_(len(ws.relationships), 2)
        eq_("rId2", ws.cell('A2').hyperlink_rel_id)
        eq_("rId2", ws.relationships[1].id)
        eq_("http://test2.com", ws.relationships[1].target)
        eq_("External", ws.relationships[1].target_mode)

    @raises(ValueError)
    def test_bad_relationship_type(self):
        rel = Relationship('bad_type')

    def test_append_list(self):
        ws = Worksheet(self.wb)

        ws.append(['This is A1', 'This is B1'])

        eq_('This is A1', ws.cell('A1').value)
        eq_('This is B1', ws.cell('B1').value)

    def test_append_dict_letter(self):
        ws = Worksheet(self.wb)

        ws.append({'A' : 'This is A1', 'C' : 'This is C1'})

        eq_('This is A1', ws.cell('A1').value)
        eq_('This is C1', ws.cell('C1').value)

    def test_append_dict_index(self):
        ws = Worksheet(self.wb)

        ws.append({0 : 'This is A1', 2 : 'This is C1'})

        eq_('This is A1', ws.cell('A1').value)
        eq_('This is C1', ws.cell('C1').value)

    @raises(TypeError)
    def test_bad_append(self):
        ws = Worksheet(self.wb)
        ws.append("test")

    def test_append_2d_list(self):

        ws = Worksheet(self.wb)

        ws.append(['This is A1', 'This is B1'])
        ws.append(['This is A2', 'This is B2'])

        vals = ws.range('A1:B2')

        eq_((('This is A1', 'This is B1'),
             ('This is A2', 'This is B2'),), flatten(vals))

    def test_rows(self):
    
        ws = Worksheet(self.wb)

        ws.cell('A1').value = 'first'
        ws.cell('C9').value = 'last'

        rows = ws.rows

        eq_(len(rows), 9)

        eq_(rows[0][0].value, 'first')
        eq_(rows[-1][-1].value, 'last')

    def test_cols(self):

        ws = Worksheet(self.wb)

        ws.cell('A1').value = 'first'
        ws.cell('C9').value = 'last'

        cols = ws.columns

        eq_(len(cols), 3)

        eq_(cols[0][0].value, 'first')
        eq_(cols[-1][-1].value, 'last')

    def test_auto_filter(self):
        ws = Worksheet(self.wb)
        ws.auto_filter = ws.range('a1:f1')
        assert ws.auto_filter == 'A1:F1'

        ws.auto_filter = ''
        assert ws.auto_filter is None

        ws.auto_filter = 'c1:g9'
        assert ws.auto_filter == 'C1:G9'

    def test_freeze(self):
        ws = Worksheet(self.wb)
        ws.freeze_panes = ws.cell('b2')
        assert ws.freeze_panes == 'B2'

        ws.freeze_panes = ''
        assert ws.freeze_panes is None

        ws.freeze_panes = 'c5'
        assert ws.freeze_panes == 'C5'

        ws.freeze_panes = ws.cell('A1')
        assert ws.freeze_panes is None

