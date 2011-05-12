# file openpyxl/tests/test_read.py

# Copyright (c) 2010 openpyxl
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# @license: http://www.opensource.org/licenses/mit-license.php
# @author: Eric Gazoni

# Python stdlib imports
from __future__ import with_statement
import os.path

# 3rd party imports
from nose.tools import eq_, raises

# package imports
from openpyxl.tests.helper import DATADIR
from openpyxl.worksheet import Worksheet
from openpyxl.workbook import Workbook
from openpyxl.style import NumberFormat, Style
from openpyxl.reader.worksheet import read_worksheet, read_dimension
from openpyxl.reader.excel import load_workbook
from openpyxl.shared.exc import InvalidFileException


def test_read_standalone_worksheet():

    class DummyWb(object):

        def get_sheet_by_name(self, value):
            return None

    path = os.path.join(DATADIR, 'reader', 'sheet2.xml')
    with open(path) as handle:
        ws = read_worksheet(handle.read(), DummyWb(),
                'Sheet 2', {1: 'hello'}, {1: Style()})
    assert isinstance(ws, Worksheet)
    eq_(ws.cell('G5').value, 'hello')
    eq_(ws.cell('D30').value, 30)
    eq_(ws.cell('K9').value, 0.09)


def test_read_standard_workbook():
    path = os.path.join(DATADIR, 'genuine', 'empty.xlsx')
    wb = load_workbook(path)
    assert isinstance(wb, Workbook)

def test_read_standard_workbook_from_fileobj():
    path = os.path.join(DATADIR, 'genuine', 'empty.xlsx')
    fo = open(path, mode = 'rb')
    wb = load_workbook(fo)
    assert isinstance(wb, Workbook)

def test_read_worksheet():
    path = os.path.join(DATADIR, 'genuine', 'empty.xlsx')
    wb = load_workbook(path)
    sheet2 = wb.get_sheet_by_name('Sheet2 - Numbers')
    assert isinstance(sheet2, Worksheet)
    eq_('This is cell G5', sheet2.cell('G5').value)
    eq_(18, sheet2.cell('D18').value)


def test_read_nostring_workbook():
    genuine_wb = os.path.join(DATADIR, 'genuine', 'empty-no-string.xlsx')
    wb = load_workbook(genuine_wb)
    assert isinstance(wb, Workbook)

@raises(InvalidFileException)
def test_read_empty_file():

    null_file = os.path.join(DATADIR, 'reader', 'null_file.xlsx')
    wb = load_workbook(null_file)

@raises(InvalidFileException)
def test_read_empty_archive():

    null_file = os.path.join(DATADIR, 'reader', 'null_archive.xlsx')
    wb = load_workbook(null_file)

def test_read_dimension():

    path = os.path.join(DATADIR, 'reader', 'sheet2.xml')

    with open(path) as handle:

        dimension = read_dimension(xml_source = handle.read())

    eq_(('D', 1, 'K', 30), dimension)

class TestReadWorkbookWithStyles(object):

    @classmethod
    def setup_class(cls):
        cls.genuine_wb = os.path.join(DATADIR, 'genuine', \
                'empty-with-styles.xlsx')
        wb = load_workbook(cls.genuine_wb)
        cls.ws = wb.get_sheet_by_name('Sheet1')

    def test_read_general_style(self):
        eq_(self.ws.cell('A1').style.number_format.format_code,
                NumberFormat.FORMAT_GENERAL)

    def test_read_date_style(self):
        eq_(self.ws.cell('A2').style.number_format.format_code,
                NumberFormat.FORMAT_DATE_XLSX14)

    def test_read_number_style(self):
        eq_(self.ws.cell('A3').style.number_format.format_code,
                NumberFormat.FORMAT_NUMBER_00)

    def test_read_time_style(self):
        eq_(self.ws.cell('A4').style.number_format.format_code,
                NumberFormat.FORMAT_DATE_TIME3)

    def test_read_percentage_style(self):
        eq_(self.ws.cell('A5').style.number_format.format_code,
                NumberFormat.FORMAT_PERCENTAGE_00)
