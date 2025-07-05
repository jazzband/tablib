#!/usr/bin/env python
"""Tests for Tablib."""

import datetime as dt
import doctest
import json
import pickle
import re
import tempfile
import unittest
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4

import xlrd
from odf import opendocument, table
from openpyxl.reader.excel import load_workbook

import tablib
from tablib.core import Row, detect_format
from tablib.exceptions import UnsupportedFormat
from tablib.formats import registry

try:
    import pandas
except ImportError:  # pragma: no cover
    pandas = None


class BaseTestCase(unittest.TestCase):
    def setUp(self):
        """Create simple data set with headers."""

        global data, book

        data = tablib.Dataset()
        book = tablib.Databook()

        self.headers = ('first_name', 'last_name', 'gpa')
        self.john = ('John', 'Adams', 90)
        self.george = ('George', 'Washington', 67)
        self.tom = ('Thomas', 'Jefferson', 50)

        self.founders = tablib.Dataset(headers=self.headers, title='Founders')
        self.founders.append(self.john)
        self.founders.append(self.george)
        self.founders.append(self.tom)


class TablibTestCase(BaseTestCase):
    """Tablib test cases."""

    def _test_export_data_in_all_formats(self, dataset, exclude=()):
        all_formats = [
            'json', 'yaml', 'csv', 'tsv', 'xls', 'xlsx', 'ods', 'html', 'jira',
            'latex', 'df', 'rst',
        ]
        for format_ in all_formats:
            if format_ in exclude or (format_ == 'df' and pandas is None):
                continue
            dataset.export(format_)

    def test_unknown_format(self):
        with self.assertRaises(UnsupportedFormat):
            data.export('??')
        # A known format but uninstalled
        saved_registry = registry._formats.copy()
        try:
            del registry._formats['ods']
            msg = (r"The 'ods' format is not available. You may want to install the "
                   "odfpy package \\(or `pip install \"tablib\\[ods\\]\"`\\).")
            with self.assertRaisesRegex(UnsupportedFormat, msg):
                data.export('ods')
        finally:
            registry._formats = saved_registry

    def test_empty_append(self):
        """Verify append() correctly adds tuple with no headers."""
        new_row = (1, 2, 3)
        data.append(new_row)

        # Verify width/data
        self.assertEqual(data.width, len(new_row))
        self.assertEqual(data[0], new_row)

    def test_empty_append_with_headers(self):
        """Verify append() correctly detects mismatch of number of
        headers and data.
        """
        data.headers = ['first', 'second']
        new_row = (1, 2, 3, 4)

        self.assertRaises(tablib.InvalidDimensions, data.append, new_row)

    def test_set_headers_with_incorrect_dimension(self):
        """Verify headers correctly detects mismatch of number of
        headers and data.
        """

        data.append(self.john)

        def set_header_callable():
            data.headers = ['first_name']

        self.assertRaises(tablib.InvalidDimensions, set_header_callable)

    def test_add_column(self):
        """Verify adding column works with/without headers."""

        data.append(['kenneth'])
        data.append(['bessie'])

        new_col = ['reitz', 'monke']

        data.append_col(new_col)

        self.assertEqual(data[0], ('kenneth', 'reitz'))
        self.assertEqual(data.width, 2)

        # With Headers
        data.headers = ('fname', 'lname')
        age_col = [21, 22]
        data.append_col(age_col, header='age')
        size_col = [1.65, 1.86]
        data.insert_col(1, size_col, header='size')

        self.assertEqual(data['age'], age_col)
        self.assertEqual(data['size'], size_col)

    def test_add_column_no_data_with_headers(self):
        """Verify adding empty column when dataset has only headers."""
        data.headers = ('fname', 'lname')
        data.insert_col(1, [], header='size')
        self.assertEqual(data.headers, ['fname', 'size', 'lname'])

    def test_add_column_no_data_no_headers(self):
        """Verify adding new column with no headers."""

        new_col = ('reitz', 'monke')

        data.append_col(new_col)

        self.assertEqual(data[0], (new_col[0],))
        self.assertEqual(data.width, 1)
        self.assertEqual(data.height, len(new_col))

    def test_add_column_with_header_ignored(self):
        """Verify append_col() ignores the header if data.headers has
        not previously been set
        """

        new_col = ('reitz', 'monke')

        data.append_col(new_col, header='first_name')

        self.assertEqual(data[0], (new_col[0],))
        self.assertEqual(data.width, 1)
        self.assertEqual(data.height, len(new_col))
        self.assertEqual(data.headers, None)

    def test_add_column_with_header_and_headers_only_exist(self):
        """Verify append_col() with header correctly detects mismatch when
        headers exist but there is no existing row data
        """

        data.headers = ['first_name']
        # no data

        new_col = 'allen'

        def append_col_callable():
            data.append_col(new_col, header='middle_name')

        self.assertRaises(tablib.InvalidDimensions, append_col_callable)

    def test_add_column_with_header_and_data_exists(self):
        """Verify append_col() works when headers and rows exists"""

        data.headers = self.headers
        data.append(self.john)

        new_col = [10]

        data.append_col(new_col, header='age')

        self.assertEqual(data.height, 1)
        self.assertEqual(data.width, len(self.john) + 1)
        self.assertEqual(data['age'], new_col)
        self.assertEqual(len(data.headers), len(self.headers) + 1)

    def test_add_callable_column(self):
        """Verify adding column with values specified as callable."""

        def new_col(row):
            return row[0]

        def initials(row):
            return f"{row[0][0]}{row[1][0]}"

        self.founders.append_col(new_col, header='first_again')
        self.founders.append_col(initials, header='initials')

        # A new row can still be appended, and the dynamic column value generated.
        self.founders.append(('Some', 'One', 71))
        # Also acceptable when all dynamic column values are provided.
        self.founders.append(('Other', 'Second', 84, 'Other', 'OS'))

        self.assertEqual(self.founders[3], ('Some', 'One', 71, 'Some', 'SO'))
        self.assertEqual(self.founders[4], ('Other', 'Second', 84, 'Other', 'OS'))
        self.assertEqual(
            self.founders['first_again'],
            ['John', 'George', 'Thomas', 'Some', 'Other']
        )
        self.assertEqual(
            self.founders['initials'],
            ['JA', 'GW', 'TJ', 'SO', 'OS']
        )

        # However only partial dynamic values provided is not accepted.
        with self.assertRaises(tablib.InvalidDimensions):
            self.founders.append(('Should', 'Crash', 60, 'Partial'))

        # Add a new row after dynamic column deletion
        del self.founders['first_again']
        self.founders.append(('After', 'Deletion', 75))
        self.assertEqual(
            self.founders['initials'],
            ['JA', 'GW', 'TJ', 'SO', 'OS', 'AD']
        )

    def test_header_slicing(self):
        """Verify slicing by headers."""

        self.assertEqual(self.founders['first_name'],
                         [self.john[0], self.george[0], self.tom[0]])

        self.assertEqual(self.founders['last_name'],
                         [self.john[1], self.george[1], self.tom[1]])

        self.assertEqual(self.founders['gpa'],
                         [self.john[2], self.george[2], self.tom[2]])

    def test_get(self):
        """Verify getting rows by index"""

        self.assertEqual(self.founders.get(0), self.john)
        self.assertEqual(self.founders.get(1), self.george)
        self.assertEqual(self.founders.get(2), self.tom)

        self.assertEqual(self.founders.get(-1), self.tom)
        self.assertEqual(self.founders.get(-2), self.george)
        self.assertEqual(self.founders.get(-3), self.john)

        with self.assertRaises(IndexError):
            self.founders.get(3)

        with self.assertRaises(TypeError):
            self.founders.get('first_name')

    def test_get_col(self):
        """Verify getting columns by index"""

        self.assertEqual(
            self.founders.get_col(list(self.headers).index('first_name')),
            [self.john[0], self.george[0], self.tom[0]])

        self.assertEqual(
            self.founders.get_col(list(self.headers).index('last_name')),
            [self.john[1], self.george[1], self.tom[1]])

        self.assertEqual(
            self.founders.get_col(list(self.headers).index('gpa')),
            [self.john[2], self.george[2], self.tom[2]])

    def test_data_slicing(self):
        """Verify slicing by data."""

        # Slice individual rows
        self.assertEqual(self.founders[0], self.john)
        self.assertEqual(self.founders[:1], [self.john])
        self.assertEqual(self.founders[1:2], [self.george])
        self.assertEqual(self.founders[-1], self.tom)
        self.assertEqual(self.founders[3:], [])

        # Slice multiple rows
        self.assertEqual(self.founders[:], [self.john, self.george, self.tom])
        self.assertEqual(self.founders[0:2], [self.john, self.george])
        self.assertEqual(self.founders[1:3], [self.george, self.tom])
        self.assertEqual(self.founders[2:], [self.tom])

    def test_row_slicing(self):
        """Verify Row slicing. Issue #184."""

        john = Row(self.john)

        self.assertEqual(john[:], list(self.john[:]))
        self.assertEqual(john[0:], list(self.john[0:]))
        self.assertEqual(john[:2], list(self.john[:2]))
        self.assertEqual(john[0:2], list(self.john[0:2]))
        self.assertEqual(john[0:-1], list(self.john[0:-1]))

    def test_delete(self):
        """Verify deleting from dataset works."""

        # Delete from front of object
        del self.founders[0]
        self.assertEqual(self.founders[:], [self.george, self.tom])

        # Verify dimensions, width should NOT change
        self.assertEqual(self.founders.height, 2)
        self.assertEqual(self.founders.width, 3)

        # Delete from back of object
        del self.founders[1]
        self.assertEqual(self.founders[:], [self.george])

        # Verify dimensions, width should NOT change
        self.assertEqual(self.founders.height, 1)
        self.assertEqual(self.founders.width, 3)

        # Delete from invalid index
        self.assertRaises(IndexError, self.founders.__delitem__, 3)

    def test_str_no_columns(self):
        d = tablib.Dataset(['a', 1], ['b', 2], ['c', 3])
        output = f'{d}'

        self.assertEqual(output.splitlines(), [
            'a|1',
            'b|2',
            'c|3'
        ])

    def test_unicode_append(self):
        """Passes in a single unicode character and exports."""

        new_row = ('å', 'é')

        data.append(new_row)
        self._test_export_data_in_all_formats(data)

    def test_datetime_append(self):
        """Passes in a single datetime and a single date and exports."""

        new_row = (
            dt.datetime.now(),
            dt.datetime.today(),
        )

        data.append(new_row)
        self._test_export_data_in_all_formats(data)

    def test_separator_append(self):
        for _ in range(3):
            data.append_separator('foobar')
            for _ in range(5):
                data.append(['asdf', 'asdf', 'asdf'])
        self._test_export_data_in_all_formats(data)

    def test_book_export_no_exceptions(self):
        """Test that various exports don't error out."""

        book = tablib.Databook()
        book.add_sheet(data)
        # These formats don't implement the book abstraction.
        unsupported = ['csv', 'tsv', 'jira', 'latex', 'df']
        self._test_export_data_in_all_formats(book, exclude=unsupported)

    def test_book_unsupported_loading(self):
        with self.assertRaises(UnsupportedFormat):
            tablib.Databook().load('Any stream', 'csv')

    def test_book_unsupported_export(self):
        book = tablib.Databook().load(
            '[{"title": "first", "data": [{"first_name": "John"}]}]',
            'json',
        )
        with self.assertRaises(UnsupportedFormat):
            book.export('csv')

    def test_book_import_from_file(self):
        xlsx_source = Path(__file__).parent / 'files' / 'founders.xlsx'
        with xlsx_source.open('rb') as fh:
            book = tablib.Databook().load(fh, 'xlsx')
        self.assertEqual(eval(book.json)[0]['title'], 'Feuille1')

    def test_dataset_import_from_file(self):
        xlsx_source = Path(__file__).parent / 'files' / 'founders.xlsx'
        with xlsx_source.open('rb') as fh:
            dset = tablib.Dataset().load(fh, 'xlsx')
        self.assertEqual(eval(dset.json)[0]['last_name'], 'Adams')

    def test_empty_file(self):
        tmp_file = tempfile.NamedTemporaryFile()
        dset = tablib.Dataset().load(tmp_file, 'yaml')
        self.assertEqual(dset.json, '[]')

    def test_auto_format_detect(self):
        """Test auto format detection."""
        # html, jira, latex, rst are export only.

        _xls = self.founders.export('xls')
        self.assertEqual(tablib.detect_format(_xls), 'xls')

        _xlsx = self.founders.export('xlsx')
        self.assertEqual(tablib.detect_format(_xlsx), 'xlsx')

        _ods = self.founders.export('ods')
        self.assertEqual(tablib.detect_format(_ods), 'ods')

        if pandas is not None:
            _df = self.founders.export('df')
            self.assertEqual(tablib.detect_format(_df), 'df')

        _yaml = '- {age: 90, first_name: John, last_name: Adams}'
        self.assertEqual(tablib.detect_format(_yaml), 'yaml')

        _json = '[{"last_name": "Adams","age": 90,"first_name": "John"}]'
        self.assertEqual(tablib.detect_format(_json), 'json')

        _csv = '1,2,3\n4,5,6\n7,8,9\n'
        self.assertEqual(tablib.detect_format(_csv), 'csv')

        _tsv = '1\t2\t3\n4\t5\t6\n7\t8\t9\n'
        self.assertEqual(tablib.detect_format(_tsv), 'tsv')

        _bunk = StringIO(
            '¡¡¡¡¡¡---///\n\n\n'
            '¡¡£™∞¢£§∞§¶•¶ª∞¶•ªº••ª–º§•†•§º¶•†¥ª–º•§ƒø¥¨©πƒø†ˆ¥ç©¨√øˆ¥≈†ƒ¥ç©ø¨çˆ¥ƒçø¶'
        )
        self.assertEqual(tablib.detect_format(_bunk), None)

    def test_transpose(self):
        """Transpose a dataset."""

        transposed_founders = self.founders.transpose()
        first_row = transposed_founders[0]
        second_row = transposed_founders[1]

        self.assertEqual(transposed_founders.headers,
                         ["first_name", "John", "George", "Thomas"])
        self.assertEqual(first_row,
                         ("last_name", "Adams", "Washington", "Jefferson"))
        self.assertEqual(second_row,
                         ("gpa", 90, 67, 50))

    def test_transpose_empty_dataset(self):
        data = tablib.Dataset()
        self.assertEqual(data.transpose(), None)

    def test_transpose_with_no_headers(self):
        data = tablib.Dataset()
        data.append(('Cat', 'Eats fish', 26))
        data.append(['Dogs like', '_balls', 48])
        data.append([73, 'people', 'sleeps'])
        dataTrans = data.transpose()
        self.assertEqual(dataTrans[0], ('Cat', 'Dogs like', 73))
        self.assertEqual(dataTrans[1], ('Eats fish', '_balls', 'people'))
        self.assertEqual(dataTrans[2], (26, 48, 'sleeps'))
        self.assertEqual(data.transpose().transpose().dict, data.dict)

    def test_transpose_multiple_headers(self):

        data = tablib.Dataset()
        data.headers = ("first_name", "last_name", "age")
        data.append(('John', 'Adams', 90))
        data.append(('George', 'Washington', 67))
        data.append(('John', 'Tyler', 71))
        self.assertEqual(data.transpose().transpose().dict, data.dict)

    def test_row_stacking(self):
        """Row stacking."""

        to_join = tablib.Dataset(headers=self.founders.headers)

        for row in self.founders:
            to_join.append(row=row)

        row_stacked = self.founders.stack(to_join)

        for column in row_stacked.headers:
            original_data = self.founders[column]
            expected_data = original_data + original_data
            self.assertEqual(row_stacked[column], expected_data)

    def test_column_stacking(self):
        """Column stacking"""

        to_join = tablib.Dataset(headers=self.founders.headers)

        for row in self.founders:
            to_join.append(row=row)

        column_stacked = self.founders.stack_cols(to_join)

        for index, row in enumerate(column_stacked):
            original_data = self.founders[index]
            expected_data = original_data + original_data
            self.assertEqual(row, expected_data)

        self.assertEqual(column_stacked[0],
                         ("John", "Adams", 90, "John", "Adams", 90))

    def test_sorting(self):
        """Sort columns."""

        sorted_data = self.founders.sort(col="first_name")
        self.assertEqual(sorted_data.title, 'Founders')

        first_row = sorted_data[0]
        second_row = sorted_data[2]
        third_row = sorted_data[1]
        expected_first = self.founders[1]
        expected_second = self.founders[2]
        expected_third = self.founders[0]

        self.assertEqual(first_row, expected_first)
        self.assertEqual(second_row, expected_second)
        self.assertEqual(third_row, expected_third)

    def test_remove_duplicates(self):
        """Unique Rows."""

        self.founders.append(self.john)
        self.founders.append(self.george)
        self.founders.append(self.tom)
        self.assertEqual(self.founders[0], self.founders[3])
        self.assertEqual(self.founders[1], self.founders[4])
        self.assertEqual(self.founders[2], self.founders[5])
        self.assertEqual(self.founders.height, 6)

        self.founders.remove_duplicates()

        self.assertEqual(self.founders[0], self.john)
        self.assertEqual(self.founders[1], self.george)
        self.assertEqual(self.founders[2], self.tom)
        self.assertEqual(self.founders.height, 3)

    def test_wipe(self):
        """Purge a dataset."""

        new_row = (1, 2, 3)
        data.append(new_row)

        # Verify width/data
        self.assertEqual(data.width, len(new_row))
        self.assertEqual(data[0], new_row)

        data.wipe()
        new_row = (1, 2, 3, 4)
        data.append(new_row)
        self.assertEqual(data.width, len(new_row))
        self.assertEqual(data[0], new_row)

    def test_subset(self):
        """Create a subset of a dataset"""

        rows = (0, 2)
        columns = ('first_name', 'gpa')

        data.headers = self.headers

        data.append(self.john)
        data.append(self.george)
        data.append(self.tom)

        # Verify data is truncated
        subset = data.subset(rows=rows, cols=columns)
        self.assertEqual(type(subset), tablib.Dataset)
        self.assertEqual(subset.headers, list(columns))
        self.assertEqual(subset._data[0].list, ['John', 90])
        self.assertEqual(subset._data[1].list, ['Thomas', 50])

    def test_formatters(self):
        """Confirm formatters are being triggered."""

        def _formatter(cell_value):
            return str(cell_value)[1:]

        self.founders.add_formatter('last_name', _formatter)

        expected = [
            {'first_name': 'John', 'last_name': 'dams', 'gpa': 90},
            {'first_name': 'George', 'last_name': 'ashington', 'gpa': 67},
            {'first_name': 'Thomas', 'last_name': 'efferson', 'gpa': 50},
        ]
        self.assertEqual(self.founders.dict, expected)
        # Test once more as the result should be the same
        self.assertEqual(self.founders.dict, expected)

    def test_formatters_all_cols(self):
        """
        Passing None as first add_formatter param apply formatter to all columns.
        """

        def _formatter(cell_value):
            return str(cell_value).upper()

        self.founders.add_formatter(None, _formatter)

        self.assertEqual(self.founders.dict, [
            {'first_name': 'JOHN', 'last_name': 'ADAMS', 'gpa': '90'},
            {'first_name': 'GEORGE', 'last_name': 'WASHINGTON', 'gpa': '67'},
            {'first_name': 'THOMAS', 'last_name': 'JEFFERSON', 'gpa': '50'},
        ])

    def test_unicode_renders_markdown_table(self):
        # add another entry to test right field width for
        # integer
        self.founders.append(('Old', 'Man', 100500))
        self.assertEqual('first_name|last_name |gpa   ', str(self.founders).split('\n')[0])

    def test_pickle_unpickle_dataset(self):
        before_pickle = self.founders.export('json')
        founders = pickle.loads(pickle.dumps(self.founders))
        self.assertEqual(founders.export('json'), before_pickle)

    def test_databook_add_sheet_accepts_only_dataset_instances(self):
        class NotDataset:
            def append(self, item):
                pass

        dataset = NotDataset()
        dataset.append(self.john)

        self.assertRaises(tablib.InvalidDatasetType, book.add_sheet, dataset)

    def test_databook_add_sheet_accepts_dataset_subclasses(self):
        class DatasetSubclass(tablib.Dataset):
            pass

        # just checking if subclass of tablib.Dataset can be added to Databook
        dataset = DatasetSubclass()
        dataset.append(self.john)
        dataset.append(self.tom)

        try:
            book.add_sheet(dataset)
        except tablib.InvalidDatasetType:
            self.fail("Subclass of tablib.Dataset should be accepted by Databook.add_sheet")

    def test_databook_formatter_support_kwargs(self):
        """Test XLSX export with formatter configuration."""
        self.founders.export('xlsx', freeze_panes=False)

    def test_databook_formatter_with_new_lines(self):
        """Test XLSX export with new line in content."""
        self.founders.append(('First\nSecond', 'Name', 42))
        self.founders.export('xlsx')

    def test_row_repr(self):
        """Row repr."""
        # Arrange
        john = Row(self.john)

        # Act
        output = str(john)

        # Assert
        self.assertEqual(output, "['John', 'Adams', 90]")

    def test_row_pickle_unpickle(self):
        """Row __setstate__ and __getstate__."""
        # Arrange
        before_pickle = Row(self.john)

        # Act
        output = pickle.loads(pickle.dumps(before_pickle))

        # Assert
        self.assertEqual(output[0], before_pickle[0])
        self.assertEqual(output[1], before_pickle[1])
        self.assertEqual(output[2], before_pickle[2])

    def test_row_lpush(self):
        """Row lpush."""
        john = Row(self.john)
        john.lpush(53)
        self.assertEqual(john.list, [53, 'John', 'Adams', 90])

    def test_row_append(self):
        """Row append."""
        john = Row(self.john)
        john.append('stuff')
        self.assertEqual(john.list, ['John', 'Adams', 90, 'stuff'])

    def test_row_contains(self):
        """Row __contains__."""
        # Arrange
        john = Row(self.john)

        # Act / Assert
        self.assertIn("John", john)

    def test_row_no_tag(self):
        """Row has_tag."""
        # Arrange
        john = Row(self.john)

        # Act / Assert
        self.assertFalse(john.has_tag("not found"))
        self.assertFalse(john.has_tag(None))

    def test_row_has_tag(self):
        """Row has_tag."""
        # Arrange
        john = Row(self.john, tags=["tag1"])

        # Act / Assert
        self.assertTrue(john.has_tag("tag1"))

    def test_row_has_tags(self):
        """Row has_tag."""
        # Arrange
        john = Row(self.john, tags=["tag1", "tag2"])

        # Act / Assert
        self.assertTrue(john.has_tag(["tag2", "tag1"]))


class HTMLTests(BaseTestCase):
    founders_html = (
        "<table>"
        "<thead>"
        "<tr><th>first_name</th><th>last_name</th><th>gpa</th></tr>"
        "</thead>"
        "<tbody>"
        "<tr><td>John</td><td>Adams</td><td>90</td></tr>"
        "<tr><td>George</td><td>Washington</td><td>67</td></tr>"
        "<tr><td>Thomas</td><td>Jefferson</td><td>50</td></tr>"
        "</tbody>"
        "</table>"
    )

    def test_html_dataset_export(self):
        """HTML export"""
        self.assertEqual(self.founders_html, self.founders.html.replace('\n', ''))

    def test_html_export_none_value(self):
        """HTML export"""

        headers = ['foo', None, 'bar']
        d = tablib.Dataset(['foø', None, 'bar'], headers=headers)
        expected = (
            "<table>"
            "<thead>"
            "<tr><th>foo</th><th></th><th>bar</th></tr>"
            "</thead>"
            "<tbody>"
            "<tr><td>foø</td><td></td><td>bar</td></tr>"
            "</tbody>"
            "</table>"
        )
        self.assertEqual(expected, d.html.replace('\n', ''))

    def test_html_databook_export(self):
        book = tablib.Databook()
        book.add_sheet(self.founders)
        book.add_sheet(self.founders)
        self.maxDiff = None
        self.assertEqual(
            book.html.replace('\n', ''),
            f"<h3>Founders</h3>{self.founders_html}<h3>Founders</h3>{self.founders_html}"
        )

    def test_html_import(self):
        data.html = self.founders_html

        self.assertEqual(['first_name', 'last_name', 'gpa'], data.headers)
        self.assertEqual([
            ('John', 'Adams', '90'),
            ('George', 'Washington', '67'),
            ('Thomas', 'Jefferson', '50'),
        ], data[:])

    def test_html_import_no_headers(self):
        data.html = """
            <table>
            <tr><td>John</td><td><i>Adams</i></td><td>90</td></tr>"
            <tr><td>George</td><td><i>Wash</i>ington</td><td>67</td></tr>"
            </table>
        """

        self.assertIsNone(data.headers)
        self.assertEqual([
            ('John', 'Adams', '90'),
            ('George', 'Washington', '67'),
        ], data[:])

    def test_html_import_no_table(self):
        html = "<html><body></body></html>"

        with self.assertRaises(ValueError) as exc:
            data.html = html
        self.assertEqual('No <table> found in input HTML', str(exc.exception))

    def test_html_import_table_id(self):
        """A table with a specific id can be targeted for import."""
        html_input = """
            <html><body>
            <table id="ignore">
              <thead><tr><th>IGNORE</th></tr></thead><tr><td>IGNORE</td></tr>
            </table>
            <table id="import">
              <thead><tr><th>first_name</th><th>last_name</th></tr></thead>
              <tr><td>John</td><td>Adams</td></tr>"
            </table>
            </html></body>
        """
        dataset = tablib.import_set(html_input, format="html", table_id="import")
        self.assertEqual(['first_name', 'last_name'], dataset.headers)
        self.assertEqual([('John', 'Adams')], dataset[:])

        # If the id is not found, an error is raised
        with self.assertRaises(ValueError) as exc:
            tablib.import_set(html_input, format="html", table_id="notfound")
        self.assertEqual('No <table> found with id="notfound" in input HTML', str(exc.exception))


class RSTTests(BaseTestCase):
    def test_rst_force_grid(self):
        data = tablib.Dataset()
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        fmt = registry.get_format('rst')
        simple = fmt.export_set(data)
        grid = fmt.export_set(data, force_grid=True)
        self.assertNotEqual(simple, grid)
        self.assertNotIn('+', simple)
        self.assertIn('+', grid)

    def test_empty_string(self):
        data = tablib.Dataset()
        data.headers = self.headers
        data.append(self.john)
        data.append(('Wendy', '', 43))
        data.append(('Esther', ' ', 31))
        self.assertEqual(
            data.export('rst'),
            '==========  =========  ===\n'
            'first_name  last_name  gpa\n'
            '==========  =========  ===\n'
            'John        Adams      90 \n'
            'Wendy                  43 \n'
            'Esther                 31 \n'
            '==========  =========  ==='
        )

    def test_rst_export_set(self):
        # Arrange
        data = tablib.Dataset()
        data.append(self.john)
        data.headers = self.headers
        fmt = registry.get_format("rst")

        # Act
        out1 = fmt.export_set(data)
        out2 = fmt.export_set_as_simple_table(data)

        # Assert
        self.assertEqual(out1, out2)
        self.assertEqual(
            out1,
            "==========  =========  ===\n"
            "first_name  last_name  gpa\n"
            "==========  =========  ===\n"
            "John        Adams      90 \n"
            "==========  =========  ===",
        )


class CSVTests(BaseTestCase):
    def test_csv_format_detect(self):
        """Test CSV format detection."""

        _csv = StringIO(
            '1,2,3\n'
            '4,5,6\n'
            '7,8,9\n'
        )
        _bunk = StringIO(
            '¡¡¡¡¡¡¡¡£™∞¢£§∞§¶•¶ª∞¶•ªº••ª–º§•†•§º¶•†¥ª–º•§ƒø¥¨©πƒø†ˆ¥ç©¨√øˆ¥≈†ƒ¥ç©ø¨çˆ¥ƒçø¶'
        )

        fmt = registry.get_format('csv')
        self.assertTrue(fmt.detect(_csv))
        self.assertFalse(fmt.detect(_bunk))

    def test_csv_import_set(self):
        """Generate and import CSV set serialization."""
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        _csv = data.csv

        data.csv = _csv

        self.assertEqual(_csv, data.csv)

    def test_csv_import_set_semicolons(self):
        """Test for proper output with semicolon separated CSV."""
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        _csv = data.get_csv(delimiter=';')

        data.set_csv(_csv, delimiter=';')

        self.assertEqual(_csv, data.get_csv(delimiter=';'))

    def test_csv_import_set_with_spaces(self):
        """Generate and import CSV set serialization when row values have
        spaces."""
        data.append(('Bill Gates', 'Microsoft'))
        data.append(('Steve Jobs', 'Apple'))
        data.headers = ('Name', 'Company')

        _csv = data.csv

        data.csv = _csv

        self.assertEqual(_csv, data.csv)

    def test_csv_import_set_semicolon_with_spaces(self):
        """Generate and import semicolon separated CSV set serialization when row values have
        spaces."""
        data.append(('Bill Gates', 'Microsoft'))
        data.append(('Steve Jobs', 'Apple'))
        data.headers = ('Name', 'Company')

        _csv = data.get_csv(delimiter=';')

        data.set_csv(_csv, delimiter=';')

        self.assertEqual(_csv, data.get_csv(delimiter=';'))

    def test_csv_import_set_with_newlines(self):
        """Generate and import CSV set serialization when row values have
        newlines."""
        data.append(('Markdown\n=======',
                     'A cool language\n\nwith paragraphs'))
        data.append(('reStructedText\n==============',
                     'Another cool language\n\nwith paragraphs'))
        data.headers = ('title', 'body')

        _csv = data.csv
        data.csv = _csv

        self.assertEqual(_csv, data.csv)

    def test_csv_import_set_commas_embedded(self):
        """Comma-separated CSV can include commas inside quoted string."""
        csv_text = (
            'id,name,description,count\r\n'
            '12,Smith,"Red, rounded",4\r\n'
        )
        data.csv = csv_text
        self.assertEqual(data[0][2], 'Red, rounded')
        self.assertEqual(data.csv, csv_text)

    def test_csv_import_set_with_unicode_str(self):
        """Import CSV set with non-ascii characters in unicode literal"""
        csv_text = (
            "id,givenname,surname,loginname,email,pref_firstname,pref_lastname\n"
            "13765,Ævar,Arnfjörð,testing,test@example.com,Ævar,Arnfjörð"
        )
        data.csv = csv_text
        self.assertEqual(data.width, 7)

    def test_csv_import_set_ragged(self):
        """Import CSV set when not all rows have the same length."""
        csv_text = (
            "H1,H2,H3\n"
            "A,B\n"
            "C,D,E\n"
            "\n"
            "F\n"
        )
        dataset = tablib.import_set(csv_text, format="csv")
        self.assertEqual(
            str(dataset),
            'H1|H2|H3\n'
            '--|--|--\n'
            'A |B |  \n'
            'C |D |E \n'
            'F |  |  '
        )

    def test_csv_import_set_skip_lines(self):
        csv_text = (
            'garbage,line\n'
            '\n'
            'id,name,description\n'
            '12,Smith,rounded\n'
        )
        dataset = tablib.import_set(csv_text, format="csv", skip_lines=2)
        self.assertEqual(dataset.headers, ['id', 'name', 'description'])

    def test_csv_import_mac_os_lf(self):
        csv_text = (
            'id,name,description\r'
            '12,Smith,rounded\r'
        )
        dataset = tablib.import_set(csv_text, format="csv")
        self.assertEqual('id,name,description\r\n12,Smith,rounded\r\n', dataset.csv)

    def test_csv_export(self):
        """Verify exporting dataset object as CSV."""

        # Build up the csv string with headers first, followed by each row
        csv = ''
        for col in self.headers:
            csv += col + ','

        csv = csv.strip(',') + '\r\n'

        for founder in self.founders:
            for col in founder:
                csv += str(col) + ','
            csv = csv.strip(',') + '\r\n'

        self.assertEqual(csv, self.founders.csv)

    def test_csv_export_options(self):
        """Exporting support csv.writer() parameters."""
        data.append(('1. a', '2. b', '3. c'))
        result = data.export('csv', delimiter=' ', quotechar='|')
        self.assertEqual(result, '|1. a| |2. b| |3. c|\r\n')

    def test_csv_stream_export(self):
        """Verify exporting dataset object as CSV from file object."""

        # Build up the csv string with headers first, followed by each row
        csv = ''
        for col in self.headers:
            csv += col + ','

        csv = csv.strip(',') + '\r\n'

        for founder in self.founders:
            for col in founder:
                csv += str(col) + ','
            csv = csv.strip(',') + '\r\n'

        frm = registry.get_format('csv')
        csv_stream = frm.export_stream_set(self.founders)
        self.assertEqual(csv, csv_stream.getvalue())

    def test_unicode_csv(self):
        """Check if unicode in csv export doesn't raise."""

        data = tablib.Dataset()

        data.append(['\xfc', '\xfd'])

        data.csv

    def test_csv_column_select(self):
        """Build up a CSV and test selecting a column"""

        data = tablib.Dataset()
        data.csv = self.founders.csv

        headers = data.headers
        self.assertIsInstance(headers[0], str)

        orig_first_name = self.founders[self.headers[0]]
        csv_first_name = data[headers[0]]
        self.assertEqual(orig_first_name, csv_first_name)

    def test_csv_column_delete(self):
        """Build up a CSV and test deleting a column"""

        data = tablib.Dataset()
        data.csv = self.founders.csv

        target_header = data.headers[0]
        self.assertIsInstance(target_header, str)

        del data[target_header]

        self.assertNotIn(target_header, data.headers)

    def test_csv_column_sort(self):
        """Build up a CSV and test sorting a column by name"""

        data = tablib.Dataset()
        data.csv = self.founders.csv

        orig_target_header = self.founders.headers[1]
        target_header = data.headers[1]

        self.founders.sort(orig_target_header)
        data.sort(target_header)

        self.assertEqual(self.founders[orig_target_header], data[target_header])

    def test_csv_formatter_support_kwargs(self):
        """Test CSV import and export with formatter configuration."""
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        expected = 'first_name;last_name;gpa\nJohn;Adams;90\nGeorge;Washington;67\n'

        kwargs = {'delimiter': ';', 'lineterminator': '\n'}
        _csv = data.export('csv', **kwargs)
        self.assertEqual(expected, _csv)

        # the import works but consider default delimiter=','
        d1 = tablib.import_set(_csv, format="csv")
        self.assertEqual(1, len(d1.headers))

        d2 = tablib.import_set(_csv, format="csv", **kwargs)
        self.assertEqual(3, len(d2.headers))


class TSVTests(BaseTestCase):
    def test_tsv_import_set(self):
        """Generate and import TSV set serialization."""
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        _tsv = data.tsv

        data.tsv = _tsv

        self.assertEqual(_tsv, data.tsv)

    def test_tsv_format_detect(self):
        """Test TSV format detection."""

        _tsv = StringIO(
            '1\t2\t3\n'
            '4\t5\t6\n'
            '7\t8\t9\n'
        )
        _bunk = StringIO(
            '¡¡¡¡¡¡¡¡£™∞¢£§∞§¶•¶ª∞¶•ªº••ª–º§•†•§º¶•†¥ª–º•§ƒø¥¨©πƒø†ˆ¥ç©¨√øˆ¥≈†ƒ¥ç©ø¨çˆ¥ƒçø¶'
        )

        fmt = registry.get_format('tsv')
        self.assertTrue(fmt.detect(_tsv))
        self.assertFalse(fmt.detect(_bunk))

    def test_tsv_export(self):
        """Verify exporting dataset object as TSV."""

        # Build up the tsv string with headers first, followed by each row
        tsv = ''
        for col in self.headers:
            tsv += col + '\t'

        tsv = tsv.strip('\t') + '\r\n'

        for founder in self.founders:
            for col in founder:
                tsv += str(col) + '\t'
            tsv = tsv.strip('\t') + '\r\n'

        self.assertEqual(tsv, self.founders.tsv)


class ODSTests(BaseTestCase):
    FORMAT_CONVERT = {
        'yearlong': '%Y',
        'monthlong': '%m',
        'daylong': '%d',
        'hourslong': '%H',
        'minuteslong': '%M',
        'secondslong': '%S',
        'secondslong0': '%S',
    }

    def test_ods_export_import_set(self):
        date = dt.date(2019, 10, 4)
        date_time = dt.datetime(2019, 10, 4, 12, 30, 8)
        time = dt.time(14, 30)
        data.append(('string', '004', 42, 21.55, Decimal('34.5'), date, time, date_time, None, ''))
        data.headers = (
            'string',
            'start0',
            'integer',
            'float',
            'decimal',
            'date',
            'time',
            'date/time',
            'None',
            'empty',
        )
        _ods = data.ods
        data.ods = _ods
        self.assertEqual(data.dict[0]['string'], 'string')
        self.assertEqual(data.dict[0]['start0'], '004')
        self.assertEqual(data.dict[0]['integer'], 42)
        self.assertEqual(data.dict[0]['float'], 21.55)
        self.assertEqual(data.dict[0]['decimal'], 34.5)
        self.assertEqual(data.dict[0]['date'], date)
        self.assertEqual(data.dict[0]['time'], time)
        self.assertEqual(data.dict[0]['date/time'], date_time)
        self.assertEqual(data.dict[0]['None'], '')
        self.assertEqual(data.dict[0]['empty'], '')

    def test_ods_export_display(self):
        """Test that exported datetime types are displayed correctly in office software"""
        date = dt.date(2019, 10, 4)
        date_time = dt.datetime(2019, 10, 4, 12, 30, 8)
        time = dt.time(14, 30)
        data.append((date, time, date_time))
        data.headers = ('date', 'time', 'date/time')
        _ods = data.ods
        ods_book = opendocument.load(BytesIO(_ods))
        styles = {style.getAttribute('name'): style for style in ods_book.styles.childNodes}
        automatic_styles = {
            style.getAttribute('name'): style.getAttribute('datastylename')
            for style in ods_book.automaticstyles.childNodes
        }

        def get_format(cell):
            style = styles[automatic_styles[cell.getAttribute('stylename')]]
            f = []
            for number in style.childNodes:
                name = number.qname[1] + ''.join(number.attributes.values())
                f.append(self.FORMAT_CONVERT.get(name, str(number)))
            return ''.join(f)

        cells = ods_book.spreadsheet.getElementsByType(table.TableRow)[1].childNodes
        self.assertEqual(str(date), str(cells[0]))
        self.assertEqual('%Y-%m-%d', get_format(cells[0]))

        self.assertEqual(str(time), str(cells[1]))
        self.assertEqual('%H:%M:%S', get_format(cells[1]))

        self.assertEqual(str(date_time), str(cells[2]))
        self.assertEqual('%Y-%m-%d %H:%M:%S', get_format(cells[2]))

    def test_ods_import_book(self):
        ods_source = Path(__file__).parent / 'files' / 'book.ods'
        with ods_source.open('rb') as fh:
            dbook = tablib.Databook().load(fh, 'ods')
        self.assertEqual(len(dbook.sheets()), 2)
        self.assertEqual(["This", "is", "a", "second", "sheet"], dbook.sheets()[1].headers)

    def test_ods_import_set_skip_lines(self):
        data.append(('garbage', 'line', ''))
        data.append(('', '', ''))
        data.append(('id', 'name', 'description'))
        _ods = data.ods
        new_data = tablib.Dataset().load(_ods, skip_lines=2)
        self.assertEqual(new_data.headers, ['id', 'name', 'description'])

    def test_ods_import_set_ragged(self):
        ods_source = Path(__file__).parent / 'files' / 'ragged.ods'
        with ods_source.open('rb') as fh:
            dataset = tablib.Dataset().load(fh, 'ods')
        self.assertEqual(dataset.pop(), (1, '', True, ''))

    def test_ods_unknown_value_type(self):
        # The ods file was trafficked to contain:
        # <table:table-cell office:value-type="unknown" calcext:value-type="string">
        ods_source = Path(__file__).parent / 'files' / 'unknown_value_type.ods'
        with ods_source.open('rb') as fh:
            dataset = tablib.Dataset().load(fh, 'ods', headers=False)
        self.assertEqual(dataset.pop(), ('abcd',))

    def test_ods_export_dates(self):
        """test against odf specification"""
        date = dt.date(2019, 10, 4)
        date_time = dt.datetime(2019, 10, 4, 12, 30, 8)
        time = dt.time(14, 30)
        data.append((date, time, date_time))
        data.headers = ('date', 'time', 'date/time')
        _ods = data.ods
        ods_book = opendocument.load(BytesIO(_ods))
        cells = ods_book.spreadsheet.getElementsByType(table.TableRow)[1].childNodes
        # date value
        self.assertEqual(cells[0].getAttribute('datevalue'), '2019-10-04')
        # time value
        duration_exp = re.compile(r"^P(?:(\d+)Y)?(?:(\d+)M)?(?:(\d+)D)?"
                                  r"(?:T(?:(\d+)H)?(?:(\d+)M)?(?:([\d.]+)S)?)?$")
        duration = duration_exp.match(cells[1].getAttribute('timevalue')).groups()
        self.assertListEqual([0, 0, 0, 14, 30, 0], [int(v or 0) for v in duration])
        # datetime value
        self.assertEqual(cells[2].getAttribute('datevalue'), '2019-10-04T12:30:08')


class XLSTests(BaseTestCase):
    def test_xls_format_detect(self):
        """Test the XLS format detection."""
        in_stream = self.founders.xls
        self.assertEqual(detect_format(in_stream), 'xls')

    def test_xls_date_import(self):
        xls_source = Path(__file__).parent / 'files' / 'dates.xls'
        with xls_source.open('rb') as fh:
            dset = tablib.Dataset().load(fh, 'xls')
        self.assertEqual(dset.dict[0]['birth_date'], dt.datetime(2015, 4, 12, 0, 0))

    def test_xlsx_import_set_skip_lines(self):
        data.append(('garbage', 'line', ''))
        data.append(('', '', ''))
        data.append(('id', 'name', 'description'))
        _xls = data.xls
        new_data = tablib.Dataset().load(_xls, skip_lines=2)
        self.assertEqual(new_data.headers, ['id', 'name', 'description'])

    def test_xls_import_with_errors(self):
        """Errors from imported files are kept as errors."""
        xls_source = Path(__file__).parent / 'files' / 'errors.xls'
        with xls_source.open('rb') as fh:
            data = tablib.Dataset().load(fh.read())
        self.assertEqual(
            data.dict[0],
            {
                'div by 0': '#DIV/0!',
                'name unknown': '#NAME?',
                'not available (formula)': '#N/A',
                'not available (static)': '#N/A',
            }
        )

    def test_book_import_from_stream(self):
        in_stream = self.founders.xls
        book = tablib.Databook().load(in_stream, 'xls')
        self.assertEqual(book.sheets()[0].title, 'Founders')

    def test_xls_export_with_dates(self):
        date = dt.date(2019, 10, 4)
        time = dt.time(14, 30)
        date_time = dt.datetime(2019, 10, 4, 12, 30, 8)
        data.append((date, time, date_time))
        data.headers = ('date', 'time', 'date/time')
        _xls = data.xls
        xls_book = xlrd.open_workbook(file_contents=_xls, formatting_info=True)
        row = xls_book.sheet_by_index(0).row(1)

        def get_format_str(cell):
            return xls_book.format_map[xls_book.xf_list[cell.xf_index].format_key].format_str

        self.assertEqual('m/d/yy', get_format_str(row[0]))
        self.assertEqual('h:mm:ss', get_format_str(row[1]))
        self.assertEqual('m/d/yy h:mm', get_format_str(row[2]))


class XLSXTests(BaseTestCase):
    def _get_width(self, data, input_arg):
        xlsx_content = data.export('xlsx', column_width=input_arg)
        wb = load_workbook(filename=BytesIO(xlsx_content))
        ws = wb.active
        return ws.column_dimensions['A'].width

    def _xlsx_cell_values_data(self, cls):
        xls_source = Path(__file__).parent / 'files' / 'xlsx_cell_values.xlsx'
        with xls_source.open('rb') as fh:
            return cls().load(fh, format='xlsx')

    def _helper_export_column_width(self, column_width):
        """check that column width adapts to value length"""

        data = self._xlsx_cell_values_data(cls=tablib.Dataset)
        width_before = self._get_width(data, column_width)
        data.append([
            'verylongvalue-verylongvalue-verylongvalue-verylongvalue-'
            'verylongvalue-verylongvalue-verylongvalue-verylongvalue',
        ])
        width_after = self._get_width(data, width_before)
        return width_before, width_after

    def _book_helper_export_column_width(self, column_width):
        """check that column width adapts to value length"""

        data = self._xlsx_cell_values_data(cls=tablib.Databook)
        width_before = self._get_width(data, column_width)
        data.sheets()[0].append([
            'verylongvalue-verylongvalue-verylongvalue-verylongvalue-'
            'verylongvalue-verylongvalue-verylongvalue-verylongvalue',
        ])
        width_after = self._get_width(data, width_before)
        return width_before, width_after

    def test_xlsx_format_detect(self):
        """Test the XLSX format detection."""
        in_stream = self.founders.xlsx
        self.assertEqual(detect_format(in_stream), 'xlsx')

    def test_xlsx_import_set(self):
        date_time = dt.datetime(2019, 10, 4, 12, 30, 8)
        data.append(('string', '004', 42, 21.55, date_time))
        data.headers = ('string', 'start0', 'integer', 'float', 'date/time')
        _xlsx = data.xlsx
        data.xlsx = _xlsx
        self.assertEqual(data.dict[0]['string'], 'string')
        self.assertEqual(data.dict[0]['start0'], '004')
        self.assertEqual(data.dict[0]['integer'], 42)
        self.assertEqual(data.dict[0]['float'], 21.55)
        self.assertEqual(data.dict[0]['date/time'], date_time)

    def test_xlsx_import_set_skip_lines(self):
        data.append(('garbage', 'line', ''))
        data.append(('', '', ''))
        data.append(('id', 'name', 'description'))
        _xlsx = data.xlsx
        new_data = tablib.Dataset().load(_xlsx, skip_lines=2)
        self.assertEqual(new_data.headers, ['id', 'name', 'description'])

    def test_xlsx_bad_chars_sheet_name(self):
        """
        Sheet names are limited to 30 chars and the following chars
        are not permitted: \\ / * ? : [ ]
        """
        _dataset = tablib.Dataset(
            title='bad name \\/*?:[]qwertyuiopasdfghjklzxcvbnm'
        )
        _xlsx = _dataset.export('xlsx')
        new_data = tablib.Dataset().load(_xlsx)
        self.assertEqual(new_data.title, 'bad name -------qwertyuiopasdfg')

        _book = tablib.Databook()
        _book.add_sheet(_dataset)
        _xlsx = _book.export('xlsx')
        new_data = tablib.Databook().load(_xlsx, 'xlsx')
        self.assertEqual(new_data.sheets()[0].title, 'bad name -------qwertyuiopasdfg')

    def test_xlsx_import_book_ragged(self):
        """Import XLSX file through databook when not all rows have the same length."""
        xlsx_source = Path(__file__).parent / 'files' / 'ragged.xlsx'
        with xlsx_source.open('rb') as fh:
            book = tablib.Databook().load(fh, 'xlsx')
        self.assertEqual(book.sheets()[0].pop(), (1.0, ''))

    def test_xlsx_import_set_ragged(self):
        """Import XLSX file through dataset when not all rows have the same length."""
        xlsx_source = Path(__file__).parent / 'files' / 'ragged.xlsx'
        with xlsx_source.open('rb') as fh:
            dataset = tablib.Dataset().load(fh, 'xlsx')
        self.assertEqual(dataset.pop(), (1.0, ''))

    def test_xlsx_wrong_char(self):
        """Bad characters are not silently ignored. We let the exception bubble up."""
        from openpyxl.utils.exceptions import IllegalCharacterError

        with self.assertRaises(IllegalCharacterError):
            data.append(('string', b'\x0cf'))
            data.xlsx

    def test_xlsx_cell_values(self):
        """Test cell values are read and not formulas"""
        xls_source = Path(__file__).parent / 'files' / 'xlsx_cell_values.xlsx'
        with xls_source.open('rb') as fh:
            data = tablib.Dataset().load(fh)
        self.assertEqual(data.headers[0], 'Hello World')

    def test_xlsx_export_set_escape_formulae(self):
        """
        Test that formulae are sanitised on export.
        """
        data.append(('=SUM(1+1)',))
        _xlsx = data.export('xlsx')

        # read back using openpyxl because tablib reads formulae as values
        wb = load_workbook(filename=BytesIO(_xlsx))
        self.assertEqual('=SUM(1+1)', wb.active['A1'].value)

        _xlsx = data.export('xlsx', escape=True)
        wb = load_workbook(filename=BytesIO(_xlsx))
        self.assertEqual('SUM(1+1)', wb.active['A1'].value)

    def test_xlsx_export_book_escape_formulae(self):
        """
        Test that formulae are sanitised on export.
        """
        data.append(('=SUM(1+1)',))
        _book = tablib.Databook()
        _book.add_sheet(data)
        _xlsx = _book.export('xlsx')

        # read back using openpyxl because tablib reads formulae as values
        wb = load_workbook(filename=BytesIO(_xlsx))
        self.assertEqual('=SUM(1+1)', wb.active['A1'].value)

        _xlsx = _book.export('xlsx', escape=True)
        wb = load_workbook(filename=BytesIO(_xlsx))
        self.assertEqual('SUM(1+1)', wb.active['A1'].value)

    def test_xlsx_export_set_escape_formulae_in_header(self):
        data.headers = ('=SUM(1+1)',)
        _xlsx = data.export('xlsx')
        wb = load_workbook(filename=BytesIO(_xlsx))
        self.assertEqual('=SUM(1+1)', wb.active['A1'].value)

        _xlsx = data.export('xlsx', escape=True)
        wb = load_workbook(filename=BytesIO(_xlsx))
        self.assertEqual('SUM(1+1)', wb.active['A1'].value)

    def test_xlsx_export_book_escape_formulae_in_header(self):
        data.headers = ('=SUM(1+1)',)
        _book = tablib.Databook()
        _book.add_sheet(data)
        _xlsx = _book.export('xlsx')
        wb = load_workbook(filename=BytesIO(_xlsx))
        self.assertEqual('=SUM(1+1)', wb.active['A1'].value)

        _xlsx = _book.export('xlsx', escape=True)
        wb = load_workbook(filename=BytesIO(_xlsx))
        self.assertEqual('SUM(1+1)', wb.active['A1'].value)

    def test_xlsx_bad_dimensions(self):
        """Test loading file with bad dimension.  Must be done with
        read_only=False."""
        xls_source = Path(__file__).parent / 'files' / 'bad_dimensions.xlsx'
        with xls_source.open('rb') as fh:
            data = tablib.Dataset().load(fh, read_only=False)
        self.assertEqual(data.height, 3)

    def test_xlsx_raise_ValueError_on_cell_write_during_export(self):
        """Test that the process handles errors which might be raised
        when calling cell setter."""
        # openpyxl does not handle array type, so will raise ValueError,
        # which results in the array being cast to string
        data.append(([1],))
        _xlsx = data.export('xlsx')
        wb = load_workbook(filename=BytesIO(_xlsx))
        self.assertEqual('[1]', wb.active['A1'].value)

    def test_xlsx_column_width_adaptive(self):
        """ Test that column width adapts to value length"""
        width_before, width_after = self._helper_export_column_width("adaptive")
        self.assertEqual(width_before, 11)
        self.assertEqual(width_after, 11)

    def test_xlsx_column_width_integer(self):
        """Test that column width changes to integer length"""
        width_before, width_after = self._helper_export_column_width(10)
        self.assertEqual(width_before, 10)
        self.assertEqual(width_after, 10)

    def test_xlsx_column_width_none(self):
        """Test that column width does not change"""
        width_before, width_after = self._helper_export_column_width(None)
        self.assertEqual(width_before, 13)
        self.assertEqual(width_after, 13)

    def test_xlsx_column_width_value_error(self):
        """Raise ValueError if column_width is not a valid input"""
        with self.assertRaises(ValueError):
            self._helper_export_column_width("invalid input")

    def test_xlsx_book_column_width_adaptive(self):
        """Test that column width adapts to value length from a databook"""
        width_before, width_after = self._book_helper_export_column_width("adaptive")
        self.assertEqual(width_before, 11)
        self.assertEqual(width_after, 11)

    def test_xlsx_book_column_width_integer(self):
        """Test that column width changes to integer length from a databook"""
        width_before, width_after = self._book_helper_export_column_width(10)
        self.assertEqual(width_before, 10)
        self.assertEqual(width_after, 10)

    def test_xlsx_book_column_width_none(self):
        """Test that column width does not change from a databook"""
        width_before, width_after = self._book_helper_export_column_width(None)
        self.assertEqual(width_before, 13)
        self.assertEqual(width_after, 13)

    def test_xlsx_book_column_width_value_error(self):
        """Raise ValueError if column_width is not a valid input for a databook"""
        with self.assertRaises(ValueError):
            self._book_helper_export_column_width("invalid input")


class JSONTests(BaseTestCase):
    def test_json_format_detect(self):
        """Test JSON format detection."""

        _json = StringIO('[{"last_name": "Adams","age": 90,"first_name": "John"}]')
        _bunk = StringIO(
            '¡¡¡¡¡¡¡¡£™∞¢£§∞§¶•¶ª∞¶•ªº••ª–º§•†•§º¶•†¥ª–º•§ƒø¥¨©πƒø†ˆ¥ç©¨√øˆ¥≈†ƒ¥ç©ø¨çˆ¥ƒçø¶'
        )

        fmt = registry.get_format('json')
        self.assertTrue(fmt.detect(_json))
        self.assertFalse(fmt.detect(_bunk))

    def test_json_import_book(self):
        """Generate and import JSON book serialization."""
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        book.add_sheet(data)
        _json = book.json

        book.json = _json

        self.assertEqual(json.loads(_json), json.loads(book.json))
        # Same with the load interface
        book2 = tablib.Databook().load(_json, None)
        self.assertEqual(json.loads(book.json), json.loads(book2.json))

    def test_json_import_set(self):
        """Generate and import JSON set serialization."""
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        _json = data.json

        data.json = _json

        self.assertEqual(json.loads(_json), json.loads(data.json))

    def test_json_export(self):
        """Verify exporting dataset object as JSON"""

        address_id = uuid4()
        headers = self.headers + ('address_id',)
        founders = tablib.Dataset(headers=headers, title='Founders')
        founders.append(('John', 'Adams', 90, address_id))
        founders.append(('名字', '李', 60, ''))
        founders_json = founders.export('json')

        expected_json = (
            '[{"first_name": "John", "last_name": "Adams", "gpa": 90, '
            f'"address_id": "{str(address_id)}"}}, {{"first_name": "名字", "last_name": "李", '
            '"gpa": 60, "address_id": ""}]'
        )

        self.assertEqual(founders_json, expected_json)

    def test_json_list_of_lists(self):
        input_json = "[[1,2],[3,4]]"
        expected_yaml = "- [1, 2]\n- [3, 4]\n"
        dset = tablib.Dataset().load(in_stream=input_json, format="json")
        self.assertEqual(dset.export("yaml"), expected_yaml)


class YAMLTests(BaseTestCase):
    def test_yaml_format_detect(self):
        """Test YAML format detection."""

        _yaml = '- {age: 90, first_name: John, last_name: Adams}'
        _tsv = 'foo\tbar'
        _bunk = (
            '¡¡¡¡¡¡---///\n\n\n¡¡£™∞¢£§∞§¶•¶ª∞¶•ªº••ª–º§•†•§º¶•†¥ª–º•§ƒø¥¨©πƒø†'
            'ˆ¥ç©¨√øˆ¥≈†ƒ¥ç©ø¨çˆ¥ƒçø¶'
        )

        fmt = registry.get_format('yaml')
        self.assertTrue(fmt.detect(_yaml))
        self.assertFalse(fmt.detect(_bunk))
        self.assertFalse(fmt.detect(_tsv))

    def test_yaml_import_book(self):
        """Generate and import YAML book serialization."""
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        book.add_sheet(data)
        _yaml = book.yaml

        book.yaml = _yaml

        self.assertEqual(_yaml, book.yaml)
        # Same with the load interface
        book2 = tablib.Databook().load(_yaml, None)
        self.assertEqual(_yaml, book2.yaml)

    def test_yaml_import_set(self):
        """Generate and import YAML set serialization."""
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        _yaml = data.yaml

        data.yaml = _yaml

        self.assertEqual(_yaml, data.yaml)

    def test_yaml_export(self):
        """YAML export"""

        self.founders.append(('名字', '李', 60))
        expected = """\
- {first_name: John, gpa: 90, last_name: Adams}
- {first_name: George, gpa: 67, last_name: Washington}
- {first_name: Thomas, gpa: 50, last_name: Jefferson}
- {first_name: 名字, gpa: 60, last_name: 李}
"""
        output = self.founders.yaml
        self.assertEqual(output, expected)

    def test_yaml_load(self):
        """ test issue 524: invalid format  """
        yaml_source = Path(__file__).parent / 'files' / 'issue_524.yaml'
        with yaml_source.open('rb') as fh:
            with self.assertRaises(UnsupportedFormat):
                tablib.Dataset().load(fh, 'yaml')


class LatexTests(BaseTestCase):
    def test_latex_export(self):
        """LaTeX export"""

        expected = """\
% Note: add \\usepackage{booktabs} to your preamble
%
\\begin{table}[!htbp]
  \\centering
  \\caption{Founders}
  \\begin{tabular}{lrr}
    \\toprule
      first\\_name & last\\_name & gpa \\\\
    \\cmidrule(r){1-1} \\cmidrule(lr){2-2} \\cmidrule(l){3-3}
      John & Adams & 90 \\\\
      George & Washington & 67 \\\\
      Thomas & Jefferson & 50 \\\\
    \\bottomrule
  \\end{tabular}
\\end{table}
"""
        output = self.founders.latex
        self.assertEqual(output, expected)

    def test_latex_export_empty_dataset(self):
        self.assertIsNotNone(tablib.Dataset().latex)

    def test_latex_export_no_headers(self):
        d = tablib.Dataset()
        d.append(('one', 'two', 'three'))
        self.assertIn('one', d.latex)

    def test_latex_export_caption(self):
        d = tablib.Dataset()
        d.append(('one', 'two', 'three'))
        self.assertNotIn('caption', d.latex)

        d.title = 'Title'
        self.assertIn('\\caption{Title}', d.latex)

    def test_latex_export_none_values(self):
        headers = ['foo', None, 'bar']
        d = tablib.Dataset(['foo', None, 'bar'], headers=headers)
        output = d.latex
        self.assertIn('foo', output)
        self.assertNotIn('None', output)

    def test_latex_escaping(self):
        d = tablib.Dataset(['~', '^'])
        output = d.latex

        self.assertNotIn('~', output)
        self.assertIn('textasciitilde', output)
        self.assertNotIn('^', output)
        self.assertIn('textasciicircum', output)


class DBFTests(BaseTestCase):
    def test_dbf_import_set(self):
        data.append(self.john)
        data.append(self.george)
        data.headers = self.headers

        _dbf = data.dbf
        data.dbf = _dbf

        # self.assertEqual(_dbf, data.dbf)
        try:
            self.assertEqual(_dbf, data.dbf)
        except AssertionError:
            index = 0
            so_far = ''
            for reg_char, data_char in zip(_dbf, data.dbf):
                so_far += chr(data_char)
                if reg_char != data_char and index not in [1, 2, 3]:
                    raise AssertionError(
                        f'Failing at char {index}: {reg_char} vs {data_char} {so_far}'
                    )
                index += 1

    def test_dbf_export_set(self):
        """Test DBF import."""
        data.append(self.john)
        data.append(self.george)
        data.append(self.tom)
        data.headers = self.headers

        _regression_dbf = (b'\x03r\x06\x06\x03\x00\x00\x00\x81\x00\xab\x00\x00'
                           b'\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00'
                           b'\x00\x00\x00FIRST_NAME\x00C\x00\x00\x00\x00P\x00\x00\x00\x00\x00'
                           b'\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00LAST_NAME\x00\x00C\x00'
                           b'\x00\x00\x00P\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00'
                           b'\x00\x00GPA\x00\x00\x00\x00\x00\x00\x00\x00N\x00\x00\x00\x00\n'
                           b'\x08\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\r'
                           )
        _regression_dbf += b' John' + (b' ' * 75)
        _regression_dbf += b' Adams' + (b' ' * 74)
        _regression_dbf += b' 90.0000000'
        _regression_dbf += b' George' + (b' ' * 73)
        _regression_dbf += b' Washington' + (b' ' * 69)
        _regression_dbf += b' 67.0000000'
        _regression_dbf += b' Thomas' + (b' ' * 73)
        _regression_dbf += b' Jefferson' + (b' ' * 70)
        _regression_dbf += b' 50.0000000'
        _regression_dbf += b'\x1a'

        # If in python3, decode regression string to binary.
        # _regression_dbf = bytes(_regression_dbf, 'utf-8')
        # _regression_dbf = _regression_dbf.replace(b'\n', b'\r')

        try:
            self.assertEqual(_regression_dbf, data.dbf)
        except AssertionError:
            index = 0
            found_so_far = ''
            for reg_char, data_char in zip(_regression_dbf, data.dbf):
                # found_so_far += chr(data_char)
                if reg_char != data_char and index not in [1, 2, 3]:
                    raise AssertionError(
                        f'Failing at char {index}: '
                        f'{reg_char} vs {data_char} (found {found_so_far})'
                    )
                index += 1

    def test_dbf_format_detect(self):
        """Test the DBF format detection."""
        _dbf = (b'\x03r\x06\x03\x03\x00\x00\x00\x81\x00\xab\x00\x00'
                b'\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00'
                b'\x00\x00\x00FIRST_NAME\x00C\x00\x00\x00\x00P\x00\x00\x00\x00\x00'
                b'\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00LAST_NAME\x00\x00C\x00'
                b'\x00\x00\x00P\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00'
                b'\x00\x00GPA\x00\x00\x00\x00\x00\x00\x00\x00N\x00\x00\x00\x00\n'
                b'\x08\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\r'
                )
        _dbf += b' John' + (b' ' * 75)
        _dbf += b' Adams' + (b' ' * 74)
        _dbf += b' 90.0000000'
        _dbf += b' George' + (b' ' * 73)
        _dbf += b' Washington' + (b' ' * 69)
        _dbf += b' 67.0000000'
        _dbf += b' Thomas' + (b' ' * 73)
        _dbf += b' Jefferson' + (b' ' * 70)
        _dbf += b' 50.0000000'
        _dbf += b'\x1a'
        _dbf = BytesIO(_dbf)

        _yaml = '- {age: 90, first_name: John, last_name: Adams}'
        _tsv = 'foo\tbar'
        _csv = '1,2,3\n4,5,6\n7,8,9\n'
        _json = '[{"last_name": "Adams","age": 90,"first_name": "John"}]'

        _bunk = (
            '¡¡¡¡¡¡¡¡£™∞¢£§∞§¶•¶ª∞¶•ªº••ª–º§•†•§º¶•†¥ª–º•§ƒø¥¨©πƒø†ˆ¥ç©¨√øˆ¥≈†ƒ¥ç©ø¨çˆ¥ƒçø¶'
        )
        fmt = registry.get_format('dbf')
        self.assertTrue(fmt.detect(_dbf))
        self.assertFalse(fmt.detect(_yaml))
        self.assertFalse(fmt.detect(_tsv))
        self.assertFalse(fmt.detect(_csv))
        self.assertFalse(fmt.detect(_json))
        self.assertFalse(fmt.detect(_bunk))


class JiraTests(BaseTestCase):
    def test_jira_export(self):
        expected = """||first_name||last_name||gpa||
|John|Adams|90|
|George|Washington|67|
|Thomas|Jefferson|50|"""
        self.assertEqual(expected, self.founders.jira)

    def test_jira_export_no_headers(self):
        self.assertEqual('|a|b|c|', tablib.Dataset(['a', 'b', 'c']).jira)

    def test_jira_export_none_and_empty_values(self):
        self.assertEqual('| | |c|', tablib.Dataset(['', None, 'c']).jira)

    def test_jira_export_empty_dataset(self):
        self.assertIsNotNone(tablib.Dataset().jira)


class DocTests(unittest.TestCase):

    def test_rst_formatter_doctests(self):
        import tablib.formats._rst
        results = doctest.testmod(tablib.formats._rst)
        self.assertEqual(results.failed, 0)


class CliTests(BaseTestCase):
    def test_cli_export_github(self):
        self.assertEqual(
            '|---|---|---|\n| a | b | c |',
            tablib.Dataset(['a', 'b', 'c']).export('cli', tablefmt='github')
        )

    def test_cli_export_simple(self):
        self.assertEqual(
            '-  -  -\na  b  c\n-  -  -',
            tablib.Dataset(['a', 'b', 'c']).export('cli', tablefmt='simple')
        )

    def test_cli_export_grid(self):
        self.assertEqual(
            '+---+---+---+\n| a | b | c |\n+---+---+---+',
            tablib.Dataset(['a', 'b', 'c']).export('cli', tablefmt='grid')
        )


class SQLFormatTests(unittest.TestCase):
    def test_sql_date_and_timestamp_literals(self):
        # ANSI SQL date and timestamp literals
        ds = tablib.Dataset(title='tbl')
        ds.headers = ['col_date', 'col_timestamp']
        ds.append([
            dt.date(2020, 1, 2),
            dt.datetime(2020, 1, 2, 3, 4, 5)
        ])
        sql = ds.export('sql')
        expected = (
            "INSERT INTO tbl (col_date,col_timestamp) VALUES "
            "(DATE '2020-01-02', TIMESTAMP '2020-01-02 03:04:05');\n"
        )
        self.assertEqual(sql, expected)

    def test_sql_microseconds_and_default_table(self):
        # Full microsecond precision and default table name 'data'
        ds = tablib.Dataset()
        ds.headers = ['ts']
        ds.append([dt.datetime(2021, 12, 31, 23, 59, 59, 123456)])
        sql = ds.export('sql')
        expected = (
            "INSERT INTO export_table (ts) VALUES "
            "(TIMESTAMP '2021-12-31 23:59:59.123456');\n"
        )
        self.assertEqual(sql, expected)

    def test_sql_regular_literals(self):
        # Test int, quoted string, decimal, bool, NULL, and multiline string
        ds = tablib.Dataset(title='t')
        ds.headers = ['i', 's', 'd', 'b', 'n', 'm', 'ml']
        ds.append([
            1,
            "O'Reilly",
            Decimal('3.14'),
            5.1,
            False,
            None,
            'Line1\nLine2'
        ])
        sql = ds.export('sql')
        expected = (
            "INSERT INTO t (i,s,d,b,n,m,ml) VALUES (1, 'O''Reilly', 3.14, 5.1, "
            "FALSE, NULL, 'Line1\nLine2');\n"
        )
        self.assertEqual(sql, expected)

    def test_sql_no_headers(self):
        # Test SQL export without headers
        ds = tablib.Dataset(title='t')
        ds.append([
            1,
            "O'Reilly",
            Decimal('3.14'),
            5.1,
            False,
            None,
            'Line1\nLine2'
        ])
        sql = ds.export('sql')
        expected = (
            "INSERT INTO t VALUES (1, 'O''Reilly', 3.14, 5.1, "
            "FALSE, NULL, 'Line1\nLine2');\n"
        )
        self.assertEqual(sql, expected)

        # Test with default table name
        ds = tablib.Dataset()
        ds.append([1, 'test'])
        sql = ds.export('sql')
        expected = "INSERT INTO export_table VALUES (1, 'test');\n"
        self.assertEqual(sql, expected)

        ds = tablib.Dataset()
        ds.append([1, 'test'])
        sql = ds.export('sql', table='schema_name.custom_table',
                        columns=['col1', 'col2'], commit=True)
        expected = ("INSERT INTO schema_name.custom_table (col1,col2)"
                    " VALUES (1, 'test');\nCOMMIT;\n")
        self.assertEqual(sql, expected)
